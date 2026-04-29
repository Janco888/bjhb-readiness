"""
BJHB Job Readiness Board Builder

Takes COOIS (components) and MB52 (stock) exports and produces a colour-coded
Excel dashboard showing which released production orders can actually start.

Algorithm: Virtual Pick Simulation
  1. Sort all released MOs by Basic Start Date (earliest first)
  2. Walk through MOs in order, simulating picking each component
  3. Stock is consumed in sequence — later MOs inherit depleted pool
  4. Classify each MO as READY / PARTIAL / NOT READY

Usage:
    python scripts/build_readiness.py
    python scripts/build_readiness.py --inputs-dir custom/path --outputs-dir out/
    python scripts/build_readiness.py --debug
"""

import argparse
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# ─── COLOUR PALETTE ──────────────────────────────────────────────────────
C = {
    "READY_BG": "C6EFCE", "READY_TXT": "375623",
    "PARTIAL_BG": "FFEB9C", "PARTIAL_TXT": "9C5700",
    "NOT_BG": "FFC7CE", "NOT_TXT": "C00000",
    "DARK": "1F3864", "WHITE": "FFFFFF",
    "SUBHDR": "2E75B6",
    "HEADER": "1F3864",
    "GREY": "F2F2F2", "GREY_TXT": "595959",
    "ROW_ALT": "F7F9FC",
}
READ_COLOR = {
    "READY": (C["READY_BG"], C["READY_TXT"]),
    "PARTIAL": (C["PARTIAL_BG"], C["PARTIAL_TXT"]),
    "NOT READY": (C["NOT_BG"], C["NOT_TXT"]),
}
COMP_COLOR = {
    "COVERED": (C["READY_BG"], C["READY_TXT"]),
    "PARTIAL": (C["PARTIAL_BG"], C["PARTIAL_TXT"]),
    "SHORT": (C["NOT_BG"], C["NOT_TXT"]),
}


# ─── STYLING HELPERS ─────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")
def bdr():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def hdr_row(ws, row, n, bg=C["HEADER"], txt=C["WHITE"], sz=10):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg); cell.font = fnt(bold=True, color=txt, size=sz)
        cell.alignment = aln("center", "center", wrap=True); cell.border = bdr()
def col_widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[get_column_letter(col)].width = w
def title_row(ws, row, text, bg, tc, sz, h, n=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg); c.font = fnt(bold=True, color=tc, size=sz)
    c.alignment = aln("center", "center", wrap=True)
    ws.row_dimensions[row].height = h


# ─── DATA LOADING ────────────────────────────────────────────────────────
def load_stock(path: Path, log: logging.Logger) -> dict[str, float]:
    """Parse MB52 hierarchical format → {material: unrestricted_qty}."""
    df = pd.read_excel(path, header=None)
    log.debug(f"MB52 raw: {len(df)} rows")

    mats = []
    i = 5
    while i < len(df):
        row = df.iloc[i]
        v0 = str(row[0]).strip() if pd.notna(row[0]) else ""
        if v0 and v0 not in ["nan", "Material", "Locat"] and pd.notna(row[9]):
            if i + 1 < len(df):
                q = df.iloc[i + 1]
                def to_num(x):
                    try: return float(x) if pd.notna(x) else 0.0
                    except (ValueError, TypeError): return 0.0
                mats.append({"Material": v0, "Unrestricted": to_num(q[5])})
        i += 1

    if not mats:
        raise ValueError("MB52 parsing returned no materials — check file format")

    df_stock = pd.DataFrame(mats)
    df_stock["Material"] = df_stock["Material"].str.strip()
    df_stock = df_stock.groupby("Material").agg(
        Unrestricted=("Unrestricted", "sum")
    ).reset_index()

    stock = dict(zip(df_stock["Material"], df_stock["Unrestricted"]))
    log.info(f"Stock loaded: {len(stock)} unique materials")
    return stock


def load_components(path: Path, today: pd.Timestamp, log: logging.Logger) -> pd.DataFrame:
    """Load COOIS and filter to relevant jobs."""
    df = pd.read_excel(path)
    log.debug(f"COOIS raw: {len(df)} rows")

    df["Material"] = df["Material"].astype(str).str.strip().replace("nan", "")
    df["Order"] = df["Order"].astype(str).str.strip()
    df["Required"] = df["Requirement Quantity"].astype(float)
    df["Withdrawn"] = df["Quantity withdrawn"].astype(float)
    df["To_Pick"] = (df["Required"] - df["Withdrawn"]).clip(lower=0)
    df["Start_Date"] = pd.to_datetime(df["Header Basic Start Date"], errors="coerce")
    df["Finish_Date"] = pd.to_datetime(df["Header Basic Finish Date"], errors="coerce")
    df["Days_to_Start"] = (df["Start_Date"] - today).dt.days
    df["Proc_Type"] = df["Procurement Type"].fillna("").astype(str).str.strip()
    df["Job_Desc"] = df["Header Material Text"].astype(str)

    if "Header SD order" in df.columns:
        df["SD_Order"] = df["Header SD order"].apply(
            lambda x: str(int(x))
            if pd.notna(x) and str(x).strip() not in ["", "nan"]
            else ""
        )
    else:
        df["SD_Order"] = ""

    # Filters
    before = len(df)
    df = df[df["To_Pick"] > 0].copy()
    df = df[df["Days_to_Start"] >= -30].copy()
    df = df[df["Material"] != ""].copy()
    log.info(
        f"Components: {before} raw → {len(df)} after filtering "
        f"(shortages, recent jobs, valid materials)"
    )
    log.info(f"Active MOs: {df['Order'].nunique()}")

    return df


def load_pos(path: Path, log: logging.Logger) -> pd.DataFrame:
    """Load ZMPO purchase order export → open PO lines by material."""
    df = pd.read_excel(path)
    df["Material"] = df["Material"].astype(str).str.strip().replace("nan", "")
    df = df[df["Material"] != ""].copy()
    df["Open_Qty"] = (df["PO-Quantity"] - df["GR-Quantity"]).clip(lower=0)
    df = df[df["Open_Qty"] > 0].copy()
    df["Delivery_Date"] = pd.to_datetime(df["Delivery Date"], errors="coerce")
    df["PO_Doc"] = df["Purchasing Document"].astype(str)
    df["Supplier_Name"] = df["Name"].astype(str).str[:30]
    log.info(f"POs loaded: {len(df)} open lines, {df['Material'].nunique()} unique materials")
    return df[["Material", "PO_Doc", "Delivery_Date", "Open_Qty", "Supplier_Name"]]


# ─── SIMULATION ──────────────────────────────────────────────────────────
def simulate_picks(
    df_comp: pd.DataFrame, stock: dict[str, float], log: logging.Logger
) -> pd.DataFrame:
    """
    Walk through MOs in start-date order, consuming stock as we go.
    Each MO's components either get fully picked (COVERED), partially picked
    (PARTIAL), or nothing (SHORT).
    """
    df_sorted = df_comp.sort_values(["Start_Date", "Order"]).copy().reset_index(drop=True)
    remaining = stock.copy()
    results = []

    for _, row in df_sorted.iterrows():
        mat = row["Material"]
        need = row["To_Pick"]
        available = remaining.get(mat, 0)
        allocated = min(need, available)
        short = need - allocated

        if allocated > 0:
            remaining[mat] = available - allocated

        if short <= 0.001:
            status = "COVERED"
        elif allocated > 0:
            status = "PARTIAL"
        else:
            status = "SHORT"

        results.append({
            "Order": row["Order"],
            "Job_Desc": row["Job_Desc"],
            "SD_Order": row["SD_Order"],
            "Start_Date": row["Start_Date"],
            "Finish_Date": row["Finish_Date"],
            "Days_to_Start": row["Days_to_Start"],
            "Material": mat,
            "Material_Desc": row["Material Description"],
            "Proc_Type": row["Proc_Type"],
            "Required": row["Required"],
            "Withdrawn": row["Withdrawn"],
            "To_Pick": need,
            "Stock_At_Turn": available + allocated,
            "Allocated": allocated,
            "Short_Qty": short,
            "Component_Status": status,
        })

    df_res = pd.DataFrame(results)
    log.info(f"Simulation complete: {df_res['Component_Status'].value_counts().to_dict()}")
    return df_res


def annotate_with_pos(
    df_comp: pd.DataFrame, df_pos: Optional[pd.DataFrame], log: logging.Logger
) -> pd.DataFrame:
    """Join open PO data to short components. Safe to call with df_pos=None."""
    empty_cols = {"PO_Doc": None, "PO_Delivery_Date": pd.NaT, "PO_Open_Qty": None}
    if df_pos is None:
        return df_comp.assign(**empty_cols)

    po_summary = (
        df_pos.sort_values("Delivery_Date")
        .groupby("Material", as_index=False)
        .agg(
            PO_Doc=("PO_Doc", "first"),
            PO_Delivery_Date=("Delivery_Date", "min"),
            PO_Open_Qty=("Open_Qty", "sum"),
        )
    )
    df_out = df_comp.merge(po_summary, on="Material", how="left")
    short_mask = df_out["Component_Status"] != "COVERED"
    with_po = (short_mask & df_out["PO_Delivery_Date"].notna()).sum()
    without_po = (short_mask & df_out["PO_Delivery_Date"].isna()).sum()
    log.info(f"Short components: {with_po} have open PO, {without_po} have no PO")
    return df_out


def aggregate_to_jobs(df_comp: pd.DataFrame, log: logging.Logger) -> pd.DataFrame:
    """Roll component outcomes up to job level with READY/PARTIAL/NOT READY."""
    def job_status(grp):
        total = len(grp)
        covered = (grp["Component_Status"] == "COVERED").sum()
        short = (grp["Component_Status"] == "SHORT").sum()

        # Readiness scored on purchased/external components only (Proc_Type != E).
        # Internal parts (E) are workshop-made — their availability is tracked
        # separately in Internal_Short but does not affect the job colour.
        ext = grp[grp["Proc_Type"] != "E"]
        ext_short = (ext["Component_Status"] != "COVERED").sum()

        if ext_short == 0:
            overall = "READY"
        elif ext_short <= 2:
            overall = "PARTIAL"
        else:
            overall = "NOT READY"

        purchased_short = (
            (grp["Proc_Type"] == "F") & (grp["Component_Status"] != "COVERED")
        ).sum()
        internal_short = (
            (grp["Proc_Type"] == "E") & (grp["Component_Status"] != "COVERED")
        ).sum()

        ext_short_rows = ext[ext["Component_Status"] != "COVERED"]
        short_with_po = ext_short_rows["PO_Delivery_Date"].notna().sum() if "PO_Delivery_Date" in ext_short_rows.columns else 0
        earliest_po = ext_short_rows["PO_Delivery_Date"].min() if "PO_Delivery_Date" in ext_short_rows.columns else pd.NaT

        return pd.Series({
            "Job_Desc": grp["Job_Desc"].iloc[0],
            "SD_Order": grp["SD_Order"].iloc[0],
            "Start_Date": grp["Start_Date"].iloc[0],
            "Finish_Date": grp["Finish_Date"].iloc[0],
            "Days_to_Start": grp["Days_to_Start"].iloc[0],
            "Total_Components": total,
            "Components_Ready": covered,
            "Components_Short": short,
            "Purchased_Short": purchased_short,
            "Internal_Short": internal_short,
            "Total_Short_Qty": grp["Short_Qty"].sum(),
            "Readiness": overall,
            "Short_with_PO": int(short_with_po),
            "Earliest_PO_Due": earliest_po,
        })

    df_jobs = df_comp.groupby("Order").apply(job_status, include_groups=False).reset_index()
    rank = {"NOT READY": 0, "PARTIAL": 1, "READY": 2}
    df_jobs["_rank"] = df_jobs["Readiness"].map(rank)
    df_jobs = df_jobs.sort_values(["_rank", "Start_Date"]).drop("_rank", axis=1).reset_index(drop=True)

    counts = df_jobs["Readiness"].value_counts().to_dict()
    log.info(f"Jobs: {counts}")
    return df_jobs


# ─── WORKBOOK BUILD ──────────────────────────────────────────────────────
def build_readiness_board(ws, df_jobs, today_str):
    ws.sheet_view.showGridLines = False
    title_row(
        ws, 1,
        f"JOB READINESS BOARD  |  {today_str}  |  Can these jobs start?",
        C["DARK"], C["WHITE"], 16, 36, 14,
    )
    title_row(
        ws, 2,
        "Virtual pick simulation — jobs ranked by start date, stock consumed in order. "
        "Earlier jobs deplete the pool, later jobs see what's left.",
        C["SUBHDR"], C["WHITE"], 10, 22, 14,
    )
    ws.row_dimensions[3].height = 8

    n_ready = int((df_jobs["Readiness"] == "READY").sum())
    n_partial = int((df_jobs["Readiness"] == "PARTIAL").sum())
    n_not = int((df_jobs["Readiness"] == "NOT READY").sum())

    kpis = [
        (n_ready, "READY to start", C["READY_BG"], C["READY_TXT"]),
        (n_partial, "PARTIAL - review", C["PARTIAL_BG"], C["PARTIAL_TXT"]),
        (n_not, "NOT READY - hold", C["NOT_BG"], C["NOT_TXT"]),
        (len(df_jobs), "Total Jobs", C["DARK"], C["WHITE"]),
    ]
    col = 1
    for val, label, bg, tc in kpis:
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 2)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 2)
        cv = ws.cell(row=4, column=col, value=val)
        cl = ws.cell(row=5, column=col, value=label)
        cv.fill = fill(bg); cl.fill = fill(bg)
        cv.font = fnt(bold=True, color=tc, size=32)
        cl.font = fnt(bold=True, color=tc, size=11)
        cv.alignment = aln("center", "center"); cl.alignment = aln("center", "center")
        ws.row_dimensions[4].height = 60; ws.row_dimensions[5].height = 28
        col += 3

    ws.row_dimensions[6].height = 8
    title_row(
        ws, 7,
        "ALL JOBS - Sorted worst first, then by Start Date  |  Filter column L to focus",
        C["SUBHDR"], C["WHITE"], 10, 22, 14,
    )

    headers = [
        "MO Number", "Job Description", "SD Order", "Start Date", "Days to Start",
        "Finish Date", "Total Parts", "Parts Ready", "Parts Short",
        "Purchased Short", "Internal Short", "READINESS",
        "Short w/ PO", "Earliest PO Due",
    ]
    hdr_row(ws, 8, 14)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=8, column=ci).value = h
    ws.row_dimensions[8].height = 30
    ws.freeze_panes = "A9"

    for ri, (_, row) in enumerate(df_jobs.iterrows(), 9):
        st = row["Readiness"]
        bg, tc = READ_COLOR[st]
        ws.cell(row=ri, column=1, value=row["Order"])
        ws.cell(row=ri, column=2, value=str(row["Job_Desc"])[:45])
        ws.cell(row=ri, column=3, value=row["SD_Order"])
        c_sd = ws.cell(
            row=ri, column=4,
            value=row["Start_Date"].date() if pd.notna(row["Start_Date"]) else None,
        )
        c_sd.number_format = "DD-MMM-YYYY"
        ws.cell(row=ri, column=5, value=f'=IF(D{ri}="","",D{ri}-TODAY())')
        c_fd = ws.cell(
            row=ri, column=6,
            value=row["Finish_Date"].date() if pd.notna(row["Finish_Date"]) else None,
        )
        c_fd.number_format = "DD-MMM-YYYY"
        ws.cell(row=ri, column=7, value=int(row["Total_Components"]))
        ws.cell(row=ri, column=8, value=int(row["Components_Ready"]))
        ws.cell(row=ri, column=9, value=int(row["Components_Short"]))
        ws.cell(row=ri, column=10, value=int(row["Purchased_Short"]))
        ws.cell(row=ri, column=11, value=int(row["Internal_Short"]))
        ws.cell(row=ri, column=12, value=st)
        ws.cell(row=ri, column=13, value=int(row["Short_with_PO"]))
        c_po = ws.cell(
            row=ri, column=14,
            value=row["Earliest_PO_Due"].date() if pd.notna(row["Earliest_PO_Due"]) else None,
        )
        c_po.number_format = "DD-MMM-YYYY"

        for ci in range(1, 15):
            c = ws.cell(row=ri, column=ci)
            c.border = bdr(); c.alignment = aln("center", "center"); c.font = fnt(size=10)
            if ci == 12:
                c.fill = fill(bg); c.font = fnt(bold=True, color=tc, size=11)
            elif ri % 2 == 0:
                c.fill = fill(C["ROW_ALT"])
        ws.row_dimensions[ri].height = 18

    last = 8 + len(df_jobs)
    ws.conditional_formatting.add(
        f"E9:E{last}",
        CellIsRule(
            operator="lessThan", formula=["0"],
            fill=fill(C["NOT_BG"]), font=fnt(bold=True, color=C["NOT_TXT"], size=10),
        ),
    )
    ws.conditional_formatting.add(
        f"E9:E{last}",
        CellIsRule(
            operator="between", formula=["0", "7"],
            fill=fill(C["PARTIAL_BG"]), font=fnt(bold=True, color=C["PARTIAL_TXT"], size=10),
        ),
    )
    ws.auto_filter.ref = f"A8:N{last}"
    col_widths(ws, {1: 13, 2: 38, 3: 13, 4: 13, 5: 11, 6: 13, 7: 10, 8: 11, 9: 11, 10: 12, 11: 11, 12: 18, 13: 12, 14: 14})


def build_component_detail(ws, df_comp, df_jobs, today_str):
    ws.sheet_view.showGridLines = False
    title_row(
        ws, 1,
        f"COMPONENT BREAKDOWN  |  {today_str}  |  Why each job got its colour",
        C["DARK"], C["WHITE"], 13, 28, 13,
    )
    ws.merge_cells("A2:P2")
    c = ws["A2"]
    c.value = (
        "Filter MO Number to see why a specific job is short. "
        "'Stock at Turn' = what was available when this MO came up in the pick sequence. "
        "Earlier MOs consume stock before later ones. PO columns show the earliest open PO for short components."
    )
    c.fill = fill(C["GREY"]); c.font = fnt(italic=True, color="595959", size=10)
    c.alignment = aln("left", "center", wrap=True)
    ws.row_dimensions[2].height = 32

    headers = [
        "MO Number", "Job Description", "Start Date", "Material", "Part Description",
        "Type", "Required", "Withdrawn", "To Pick", "Stock at Turn", "Allocated", "Short Qty",
        "Component Status", "PO Number", "PO Due Date", "PO Open Qty",
    ]
    hdr_row(ws, 3, 16)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci).value = h
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = "A4"

    rank = {"NOT READY": 0, "PARTIAL": 1, "READY": 2}
    df_c = df_comp.merge(df_jobs[["Order", "Readiness"]], on="Order", how="left")
    df_c["_rank"] = df_c["Readiness"].map(rank)
    df_c = df_c.sort_values(["_rank", "Start_Date", "Order", "Component_Status"]).drop("_rank", axis=1)

    for ri, (_, row) in enumerate(df_c.iterrows(), 4):
        cst = row["Component_Status"]
        bg, tc = COMP_COLOR[cst]
        ws.cell(row=ri, column=1, value=row["Order"])
        ws.cell(row=ri, column=2, value=str(row["Job_Desc"])[:40])
        c_sd = ws.cell(
            row=ri, column=3,
            value=row["Start_Date"].date() if pd.notna(row["Start_Date"]) else None,
        )
        c_sd.number_format = "DD-MMM-YYYY"
        ws.cell(row=ri, column=4, value=str(row["Material"])[:22])
        ws.cell(row=ri, column=5, value=str(row["Material_Desc"])[:40])
        ws.cell(row=ri, column=6, value=row["Proc_Type"])
        ws.cell(row=ri, column=7, value=round(float(row["Required"]), 2))
        ws.cell(row=ri, column=8, value=round(float(row["Withdrawn"]), 2))
        ws.cell(row=ri, column=9, value=round(float(row["To_Pick"]), 2))
        ws.cell(row=ri, column=10, value=round(float(row["Stock_At_Turn"]), 2))
        ws.cell(row=ri, column=11, value=round(float(row["Allocated"]), 2))
        ws.cell(row=ri, column=12, value=round(float(row["Short_Qty"]), 2))
        ws.cell(row=ri, column=13, value=cst)

        if cst != "COVERED":
            ws.cell(row=ri, column=14, value=row.get("PO_Doc") if pd.notna(row.get("PO_Doc")) else None)
            c_pod = ws.cell(
                row=ri, column=15,
                value=row["PO_Delivery_Date"].date() if pd.notna(row.get("PO_Delivery_Date")) else None,
            )
            c_pod.number_format = "DD-MMM-YYYY"
            po_qty = row.get("PO_Open_Qty")
            ws.cell(row=ri, column=16, value=round(float(po_qty), 2) if pd.notna(po_qty) else None)

        for ci in range(1, 17):
            c = ws.cell(row=ri, column=ci)
            c.border = bdr(); c.alignment = aln("center", "center"); c.font = fnt(size=10)
            if ci == 13:
                c.fill = fill(bg); c.font = fnt(bold=True, color=tc, size=10)
            elif ri % 2 == 0:
                c.fill = fill(C["ROW_ALT"])
        ws.row_dimensions[ri].height = 16

    last = 3 + len(df_c)
    ws.auto_filter.ref = f"A3:P{last}"
    col_widths(ws, {1: 13, 2: 35, 3: 13, 4: 20, 5: 35, 6: 6, 7: 10, 8: 10, 9: 10, 10: 13, 11: 11, 12: 11, 13: 19, 14: 14, 15: 14, 16: 12})


def build_stock_ledger(ws, df_comp, stock, today_str):
    ws.sheet_view.showGridLines = False
    title_row(
        ws, 1,
        f"STOCK CONSUMPTION LEDGER  |  {today_str}  |  How starting stock gets consumed MO by MO",
        C["DARK"], C["WHITE"], 13, 28, 9,
    )
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = (
        "When MB52 shows 50 units but the dashboard says a later job is short — this is why. "
        "Materials are listed with all MOs that claim them, consumed in start-date order."
    )
    c.fill = fill(C["GREY"]); c.font = fnt(italic=True, color="595959", size=10)
    c.alignment = aln("left", "center", wrap=True)
    ws.row_dimensions[2].height = 32

    # Build ledger
    ledger = []
    unique_mats = sorted([str(m) for m in df_comp["Material"].unique() if pd.notna(m)])
    for mat in unique_mats:
        events = df_comp[df_comp["Material"] == mat].sort_values("Start_Date").copy()
        starting = stock.get(mat, 0)
        running = starting
        for _, ev in events.iterrows():
            allocated = ev["Allocated"]
            after = running - allocated
            ledger.append({
                "Material": mat,
                "Material_Desc": ev["Material_Desc"],
                "Starting_Stock": starting,
                "MO": ev["Order"],
                "Job_Desc": ev["Job_Desc"],
                "Start_Date": ev["Start_Date"],
                "Requested": ev["To_Pick"],
                "Allocated": allocated,
                "Remaining_After": after,
            })
            running = after

    df_l = pd.DataFrame(ledger)

    headers = ["Material", "Description", "Starting Stock", "MO", "Job", "Start Date",
               "Requested", "Allocated", "Remaining After"]
    hdr_row(ws, 3, 9)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci).value = h
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = "A4"

    prev_mat = None
    for ri, (_, row) in enumerate(df_l.iterrows(), 4):
        new_mat = row["Material"] != prev_mat
        prev_mat = row["Material"]

        ws.cell(row=ri, column=1, value=str(row["Material"])[:22] if new_mat else "")
        ws.cell(row=ri, column=2, value=str(row["Material_Desc"])[:35] if new_mat else "")
        ws.cell(row=ri, column=3, value=round(float(row["Starting_Stock"]), 2) if new_mat else "")
        ws.cell(row=ri, column=4, value=row["MO"])
        ws.cell(row=ri, column=5, value=str(row["Job_Desc"])[:30])
        c_sd = ws.cell(
            row=ri, column=6,
            value=row["Start_Date"].date() if pd.notna(row["Start_Date"]) else None,
        )
        c_sd.number_format = "DD-MMM-YYYY"
        ws.cell(row=ri, column=7, value=round(float(row["Requested"]), 2))
        ws.cell(row=ri, column=8, value=round(float(row["Allocated"]), 2))
        ws.cell(row=ri, column=9, value=round(float(row["Remaining_After"]), 2))

        for ci in range(1, 10):
            c = ws.cell(row=ri, column=ci)
            c.border = bdr(); c.alignment = aln("center", "center"); c.font = fnt(size=10)
            if ri % 2 == 0:
                c.fill = fill(C["ROW_ALT"])
            if ci == 8 and row["Allocated"] < row["Requested"]:
                c.fill = fill(C["NOT_BG"]); c.font = fnt(bold=True, color=C["NOT_TXT"], size=10)
        ws.row_dimensions[ri].height = 16

    last = 3 + len(df_l)
    ws.auto_filter.ref = f"A3:I{last}"
    col_widths(ws, {1: 20, 2: 32, 3: 13, 4: 13, 5: 30, 6: 13, 7: 10, 8: 10, 9: 14})


def build_how_it_works(ws):
    ws.sheet_view.showGridLines = False
    title_row(
        ws, 1,
        "JOB READINESS BOARD - HOW IT WORKS",
        C["DARK"], C["WHITE"], 14, 32, 6,
    )
    rows = [
        ("", False, 10, C["WHITE"], "000000", 6),
        ("THE QUESTION THIS DASHBOARD ANSWERS", True, 13, C["SUBHDR"], C["WHITE"], 24),
        ("For each released production order: can it start today, or will it run short when workshop tries to pick?",
         False, 11, C["WHITE"], "000000", 32),
        ("", False, 10, C["WHITE"], "000000", 6),
        ("THE KEY INSIGHT - STOCK GETS CONSUMED IN ORDER", True, 13, C["SUBHDR"], C["WHITE"], 24),
        ("MB52 shows 100 units. If Job A needs 30, Job B needs 40, Job C needs 40 - Job C is 10 short even though stock shows 100. Earlier jobs pick first.",
         False, 10, C["WHITE"], "000000", 40),
        ("", False, 10, C["WHITE"], "000000", 6),
        ("READINESS COLOURS", True, 13, C["SUBHDR"], C["WHITE"], 24),
        ("GREEN = READY - all parts allocated - release to workshop",
         True, 11, C["READY_BG"], C["READY_TXT"], 24),
        ("YELLOW = PARTIAL - few parts short - planner decides",
         True, 11, C["PARTIAL_BG"], C["PARTIAL_TXT"], 24),
        ("RED = NOT READY - significant parts missing - do not release",
         True, 11, C["NOT_BG"], C["NOT_TXT"], 24),
        ("", False, 10, C["WHITE"], "000000", 6),
        ("THE THREE DATA SHEETS", True, 13, C["SUBHDR"], C["WHITE"], 24),
        ("READINESS_BOARD - one row per job - scan this Monday morning",
         False, 11, C["WHITE"], "000000", 24),
        ("COMPONENT_DETAIL - one row per component - filter by MO to see why a job is short",
         False, 11, C["WHITE"], "000000", 24),
        ("STOCK_LEDGER - receipts proving the math - use when someone questions a result",
         False, 11, C["WHITE"], "000000", 24),
        ("", False, 10, C["WHITE"], "000000", 6),
        ("LIMITATIONS", True, 13, C["SUBHDR"], C["WHITE"], 24),
        ("INTERNAL parts (Type E) do NOT affect job colour — readiness is scored on purchased/external parts only. Internal shortages are still shown in the 'Internal Short' column for the workshop's awareness.",
         False, 10, C["PARTIAL_BG"], C["PARTIAL_TXT"], 40),
        ("Stock = unrestricted only. QI, Transit, Blocked not included.",
         False, 10, C["PARTIAL_BG"], C["PARTIAL_TXT"], 24),
        ("No PO/PR data. A 'short' job with PO arriving tomorrow is still flagged short.",
         False, 10, C["PARTIAL_BG"], C["PARTIAL_TXT"], 24),
    ]
    for ri, item in enumerate(rows, 2):
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
        c = ws.cell(row=ri, column=1, value=item[0])
        c.fill = fill(item[3]); c.font = fnt(bold=item[1], color=item[4], size=item[2])
        c.alignment = aln("left", "center", wrap=True)
        ws.row_dimensions[ri].height = item[5]
    col_widths(ws, {1: 120})


# ─── ARCHIVE ──────────────────────────────────────────────────────────
def archive_inputs(inputs_dir: Path, log: logging.Logger) -> None:
    """Move the processed inputs to archive/ with today's date prefix."""
    archive_dir = inputs_dir / "archive"
    archive_dir.mkdir(exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")

    for name in ["coois_components.xlsx", "mb52_stock.xlsx", "y00_zmpo.xlsx"]:
        src = inputs_dir / name
        if src.exists():
            dst = archive_dir / f"{date_str}_{name}"
            shutil.copy2(src, dst)
            log.info(f"Archived: {src.name} -> archive/{dst.name}")


# ─── MAIN ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Build Job Readiness Dashboard from SAP exports"
    )
    parser.add_argument(
        "--inputs-dir", type=Path,
        default=Path(__file__).parent.parent / "inputs",
        help="Directory with coois_components.xlsx and mb52_stock.xlsx",
    )
    parser.add_argument(
        "--outputs-dir", type=Path,
        default=Path(__file__).parent.parent / "outputs",
        help="Where to write the dashboard",
    )
    parser.add_argument("--debug", action="store_true", help="Verbose logging")
    parser.add_argument("--no-archive", action="store_true", help="Skip archiving input files")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )
    log = logging.getLogger("readiness")

    inputs = args.inputs_dir.resolve()
    outputs = args.outputs_dir.resolve()
    outputs.mkdir(exist_ok=True, parents=True)

    coois_path = inputs / "coois_components.xlsx"
    mb52_path = inputs / "mb52_stock.xlsx"
    po_path = inputs / "y00_zmpo.xlsx"

    if not coois_path.exists():
        log.error(f"Missing: {coois_path}"); sys.exit(1)
    if not mb52_path.exists():
        log.error(f"Missing: {mb52_path}"); sys.exit(1)

    today = pd.Timestamp.today().normalize()
    today_str = today.strftime("%d %b %Y")

    log.info("Loading data...")
    stock = load_stock(mb52_path, log)
    df_comp_raw = load_components(coois_path, today, log)

    df_pos = None
    if po_path.exists():
        log.info("Loading PO data...")
        df_pos = load_pos(po_path, log)
    else:
        log.info("No y00_zmpo.xlsx found — PO coverage columns will be blank")

    log.info("Running simulation...")
    df_comp = simulate_picks(df_comp_raw, stock, log)
    df_comp = annotate_with_pos(df_comp, df_pos, log)

    log.info("Aggregating to job level...")
    df_jobs = aggregate_to_jobs(df_comp, log)

    log.info("Building workbook...")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "READINESS_BOARD"
    build_readiness_board(ws1, df_jobs, today_str)
    build_component_detail(wb.create_sheet("COMPONENT_DETAIL"), df_comp, df_jobs, today_str)
    build_stock_ledger(wb.create_sheet("STOCK_LEDGER"), df_comp, stock, today_str)
    build_how_it_works(wb.create_sheet("HOW_IT_WORKS"))

    stamp = datetime.now().strftime("%Y-%m-%d")
    out_path = outputs / f"Job_Readiness_Board_{stamp}.xlsx"
    wb.save(out_path)
    log.info(f"Saved: {out_path}")

    if not args.no_archive:
        archive_inputs(inputs, log)

    # Summary
    print()
    print("=" * 60)
    print(f"READINESS SUMMARY — {today_str}")
    print("=" * 60)
    n_ready = int((df_jobs["Readiness"] == "READY").sum())
    n_partial = int((df_jobs["Readiness"] == "PARTIAL").sum())
    n_not = int((df_jobs["Readiness"] == "NOT READY").sum())
    print(f"  READY:     {n_ready:>3} jobs can start now")
    print(f"  PARTIAL:   {n_partial:>3} jobs have small shortages")
    print(f"  NOT READY: {n_not:>3} jobs blocked by missing parts")
    print(f"  TOTAL:     {len(df_jobs):>3}")
    print()
    print(f"Output: {out_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
