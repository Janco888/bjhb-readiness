"""
BJHB Job Readiness Board — Streamlit App

Upload COOIS + MB52 (and optionally ZMPO), click Build, download the Excel.
"""

import importlib.util
import io
import logging
import os
import pathlib
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

_br_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "build_readiness.py")
_spec = importlib.util.spec_from_file_location("build_readiness", _br_path)
assert _spec is not None and _spec.loader is not None, f"Cannot load build_readiness from {_br_path}"
_br = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_br)  # type: ignore[union-attr]

aggregate_to_jobs = _br.aggregate_to_jobs
annotate_subassembly_risk = _br.annotate_subassembly_risk
annotate_with_pos = _br.annotate_with_pos
annotate_with_prs = _br.annotate_with_prs
build_component_detail = _br.build_component_detail
build_how_it_works = _br.build_how_it_works
build_readiness_board = _br.build_readiness_board
build_stock_ledger = _br.build_stock_ledger
load_coois_header = _br.load_coois_header
load_components = _br.load_components
load_pos = _br.load_pos
load_prs = _br.load_prs
load_stock = _br.load_stock
simulate_picks = _br.simulate_picks

READINESS_BG = {"READY": "#C6EFCE", "PARTIAL": "#FFEB9C", "NOT READY": "#FFC7CE"}
COMPONENT_BG = {"COVERED": "#C6EFCE", "PARTIAL": "#FFEB9C", "SHORT": "#FFC7CE"}

_COOIS_REQUIRED = [
    "Order", "Material", "Material Description",
    "Requirement Quantity", "Quantity withdrawn", "Procurement Type",
    "Header Material Text", "Header Basic Start Date", "Header Basic Finish Date",
]
_PO_REQUIRED = [
    "Material",
]
_OPS_REQUIRED = ["Order", "Work Center", "Standard value 2"]
_PR_REQUIRED = ["Material", "Purchase Requisition"]

st.set_page_config(page_title="BJHB Job Readiness", layout="wide")

st.title("BJHB Job Readiness Board")

# ─── SIDEBAR: file uploads + build button ────────────────────────────────────
with st.sidebar:
    st.header("SAP Exports")
    st.markdown("**Required**")
    coois_file = st.file_uploader("COOIS Components", type=["xlsx"], key="coois")
    mb52_file = st.file_uploader("MB52 Stock", type=["xlsx"], key="mb52")

    with st.expander("Optional files", expanded=True):
        po_file = st.file_uploader("ZMPO Purchase Orders", type=["xlsx"], key="po",
            help="Transaction ZMPO — open POs. Enables Supply Risk column.")
        pr_file = st.file_uploader("ME5A Purchase Requisitions", type=["xlsx"], key="pr",
            help="Transaction ME5A — open PRs. Enables PR Only status on shorts.")
        ops_file = st.file_uploader("COOIS – Operations view", type=["xlsx"], key="ops",
            help="COOIS list type = Operations. Needs Work Center + Standard value 2. Enables Production Load tab.")
        header_file = st.file_uploader("COOIS – Header view", type=["xlsx"], key="header",
            help="COOIS list type = Header. Needs Superior Order column. Enables Sub-assembly Risk column.")

    st.divider()

    ready_to_run = bool(coois_file and mb52_file)
    _build_clicked = st.button("Build Dashboard", type="primary", disabled=not ready_to_run, use_container_width=True)

    if not ready_to_run and "df_jobs" not in st.session_state:
        st.caption("Upload COOIS Components and MB52 Stock to enable the build.")

    with st.expander("What goes in each file?"):
        st.markdown(
            "**COOIS Components** *(required)*\n"
            "Transaction: `COOIS` → Component view\n"
            "Filter: Released orders, cover next 4–8 weeks\n\n"
            "**MB52 Stock** *(required)*\n"
            "Transaction: `MB52`, standard hierarchical layout\n\n"
            "**ZMPO POs** *(optional)*\n"
            "All open POs — enables Supply Risk column\n\n"
            "**ME5A PRs** *(optional)*\n"
            "All open PRs — enables PR Only status\n\n"
            "**COOIS Operations** *(optional)*\n"
            "Transaction: `COOIS` → Operation view\n"
            "Enables Production Load tab\n\n"
            "**COOIS Header** *(optional)*\n"
            "Transaction: `COOIS` → Header view\n"
            "Enables Sub-assembly Risk column"
        )

def _build_2wk_excel(df_2wk: "pd.DataFrame", df_comp: "pd.DataFrame", today_str: str) -> bytes:
    """Build a 2-sheet Excel: job summary + full component detail for next-2-weeks jobs."""
    _HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    _HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    _HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _STATUS_FILL = {
        "READY":     PatternFill("solid", fgColor="C6EFCE"),
        "PARTIAL":   PatternFill("solid", fgColor="FFEB9C"),
        "NOT READY": PatternFill("solid", fgColor="FFC7CE"),
        "COVERED":   PatternFill("solid", fgColor="C6EFCE"),
        "SHORT":     PatternFill("solid", fgColor="FFC7CE"),
    }

    def _style_sheet(ws, status_col_idx: int | None, status_map: dict):
        for cell in ws[1]:
            cell.font = _HDR_FONT
            cell.fill = _HDR_FILL
            cell.alignment = _HDR_ALIGN
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30
        if status_col_idx:
            for row in ws.iter_rows(min_row=2):
                status_val = str(row[status_col_idx - 1].value or "")
                row_fill = status_map.get(status_val)
                if row_fill:
                    for cell in row:
                        cell.fill = row_fill
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 45)

    orders_2wk = set(df_2wk["Order"].tolist())
    comp_2wk = df_comp[df_comp["Order"].isin(orders_2wk)].copy()
    comp_2wk = comp_2wk.sort_values(["Start_Date", "Order", "Component_Status", "Material"])

    # ── Sheet 1: job summary ────────────────────────────────────────────────
    _job_col_map = {
        "Order": "MO Number", "Job_Desc": "Job Description", "SD_Order": "SD Order",
        "Start_Date": "Start Date", "Finish_Date": "Required Del. Date",
        "Readiness": "Status", "Components_Short": "Parts Short",
        "Purchased_Short": "Purchased Short", "Total_Short_Qty": "Total Short Qty",
        "Supply_Risk": "Supply Risk",
    }
    job_cols = [c for c in _job_col_map if c in df_2wk.columns]
    jobs_out = df_2wk[job_cols].rename(columns=_job_col_map).copy()
    for dc in ("Start Date", "Required Del. Date"):
        if dc in jobs_out.columns:
            jobs_out[dc] = pd.to_datetime(jobs_out[dc], errors="coerce").dt.strftime("%d %b %Y")

    # ── Sheet 2: component detail ───────────────────────────────────────────
    _comp_col_map = {
        "Order": "MO Number", "Job_Desc": "Job Description", "SD_Order": "SD Order",
        "Start_Date": "Start Date", "Finish_Date": "Required Del. Date",
        "Material": "Material", "Material_Desc": "Description", "Proc_Type": "Type",
        "Required": "Required Qty", "Withdrawn": "Withdrawn Qty",
        "To_Pick": "To Pick", "Stock_At_Turn": "Stock Available",
        "Allocated": "Allocated", "Short_Qty": "Short Qty",
        "Component_Status": "Status",
        "PO_Doc": "PO Number", "PO_Open_Qty": "PO Open Qty", "PO_Delivery_Date": "PO Delivery Date",
        "PR_Doc": "PR Number", "PR_Open_Qty": "PR Open Qty", "PR_Delivery_Date": "PR Delivery Date",
    }
    comp_cols = [c for c in _comp_col_map if c in comp_2wk.columns]
    comp_out = comp_2wk[comp_cols].rename(columns=_comp_col_map).copy()
    for dc in ("Start Date", "Required Del. Date", "PO Delivery Date", "PR Delivery Date"):
        if dc in comp_out.columns:
            comp_out[dc] = pd.to_datetime(comp_out[dc], errors="coerce").dt.strftime("%d %b %Y").where(
                pd.to_datetime(comp_out[dc], errors="coerce").notna(), ""
            )

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        jobs_out.to_excel(writer, sheet_name="Jobs", index=False)
        comp_out.to_excel(writer, sheet_name="Components", index=False)

        jobs_status_idx = (list(jobs_out.columns).index("Status") + 1) if "Status" in jobs_out.columns else None
        comp_status_idx = (list(comp_out.columns).index("Status") + 1) if "Status" in comp_out.columns else None

        _style_sheet(writer.sheets["Jobs"], jobs_status_idx, _STATUS_FILL)
        _style_sheet(writer.sheets["Components"], comp_status_idx, _STATUS_FILL)

        # Title rows above headers
        for sheet_name, title in [
            ("Jobs", f"Next 2 Weeks — Job Summary  ({today_str})"),
            ("Components", f"Next 2 Weeks — Component Detail  ({today_str})"),
        ]:
            ws = writer.sheets[sheet_name]
            ws.insert_rows(1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
            title_cell.fill = PatternFill("solid", fgColor="1F4E79")
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 20
            ws.freeze_panes = "A3"

    out.seek(0)
    return out.getvalue()


_INPUTS_DIR = pathlib.Path(os.path.dirname(os.path.abspath(__file__))) / "inputs"
_INPUTS_DIR.mkdir(exist_ok=True)
for _uf, _fname in [
    (coois_file, "coois_components.xlsx"),
    (mb52_file, "mb52_stock.xlsx"),
    (po_file, "y00_zmpo.xlsx"),
    (pr_file, "me5a_prs.xlsx"),
    (ops_file, "coois_operations.xlsx"),
    (header_file, "coois_header.xlsx"),
]:
    if _uf is not None:
        _uf.seek(0)
        _dest = _INPUTS_DIR / _fname
        try:
            _dest.write_bytes(_uf.read())
        except PermissionError:
            if not _dest.exists():
                raise
            st.warning(f"Could not update {_fname} on disk (file locked) — using existing copy.")
        _uf.seek(0)

if _build_clicked:
    log_lines = []

    class ListHandler(logging.Handler):
        def emit(self, record):
            log_lines.append(self.format(record))

    log = logging.getLogger("readiness_app")
    log.handlers.clear()
    handler = ListHandler()
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S"))
    log.addHandler(handler)
    log.setLevel(logging.INFO)
    log.propagate = False

    try:
        # ── Pre-flight validation ──────────────────────────────────────────
        val_errors = []

        coois_file.seek(0)
        try:
            _df_coois_chk = pd.read_excel(coois_file)
            missing = [c for c in _COOIS_REQUIRED if c not in _df_coois_chk.columns]
            if missing:
                val_errors.append(
                    "**COOIS file is missing required columns:**\n\n"
                    + "\n".join(f"- `{c}`" for c in missing)
                    + "\n\nMake sure you exported from COOIS in **Component view** with the full layout."
                )
            elif len(_df_coois_chk) == 0:
                val_errors.append("**COOIS file is empty.** Check your SAP filter settings.")
            coois_file.seek(0)
        except Exception as e:
            val_errors.append(f"**COOIS file could not be read:** {e}")

        mb52_file.seek(0)
        try:
            _df_mb52_chk = pd.read_excel(mb52_file, header=None, nrows=10)
            if len(_df_mb52_chk) < 3:
                val_errors.append("**MB52 file appears empty or too small.** Check your SAP export.")
            mb52_file.seek(0)
        except Exception as e:
            val_errors.append(f"**MB52 file could not be read:** {e}")

        if po_file:
            po_file.seek(0)
            try:
                _df_po_chk = pd.read_excel(po_file)
                missing_po = [c for c in _PO_REQUIRED if c not in _df_po_chk.columns]
                if missing_po:
                    val_errors.append(
                        "**ZMPO file is missing required columns:**\n\n"
                        + "\n".join(f"- `{c}`" for c in missing_po)
                        + "\n\nCheck that you exported from the ZMPO transaction with the standard layout."
                    )
                po_file.seek(0)
            except Exception as e:
                val_errors.append(f"**ZMPO file could not be read:** {e}")

        if pr_file:
            pr_file.seek(0)
            try:
                _df_pr_chk = pd.read_excel(pr_file)
                _df_pr_chk.columns = _df_pr_chk.columns.str.strip()
                missing_pr = [c for c in _PR_REQUIRED if c not in _df_pr_chk.columns]
                if missing_pr:
                    val_errors.append(
                        "**ME5A file is missing required columns:**\n\n"
                        + "\n".join(f"- `{c}`" for c in missing_pr)
                        + "\n\nMake sure your ME5A layout includes the Material and Purchase Requisition columns."
                    )
                pr_file.seek(0)
            except Exception as e:
                val_errors.append(f"**ME5A file could not be read:** {e}")

        if ops_file:
            ops_file.seek(0)
            try:
                _df_ops_chk = pd.read_excel(ops_file)
                _df_ops_chk.columns = _df_ops_chk.columns.str.strip()
                missing_ops = [c for c in _OPS_REQUIRED if c not in _df_ops_chk.columns]
                if missing_ops:
                    val_errors.append(
                        "**COOIS Operations file is missing required columns:**\n\n"
                        + "\n".join(f"- `{c}`" for c in missing_ops)
                        + "\n\nIn COOIS set the list type to **Operations** and ensure the layout includes Order, Work Center, and Standard value 2."
                    )
                ops_file.seek(0)
            except Exception as e:
                val_errors.append(f"**COOIS Operations file could not be read:** {e}")

        if val_errors:
            for err in val_errors:
                st.error(err)
            st.stop()

        with st.spinner("Loading data..."):
            today = pd.Timestamp.today().normalize()
            today_str = today.strftime("%d %b %Y")
            stock, sd_stock = load_stock(mb52_file, log)
            df_comp_raw = load_components(coois_file, today, log)
            df_pos = load_pos(po_file, log) if po_file else None
            df_prs = load_prs(pr_file, log) if pr_file else None
            df_header = load_coois_header(header_file, log) if header_file else None

            if ops_file:
                ops_file.seek(0)
                df_ops = pd.read_excel(ops_file)
                df_ops.columns = df_ops.columns.str.strip()
                df_ops["Order"] = df_ops["Order"].astype(str).str.strip().str.split(".").str[0]
                df_ops = df_ops.rename(columns={
                    "Work Center": "Work_Centre",
                    "Work center description": "Work_Centre_Desc",
                    "Standard value 2": "Planned_Hrs",
                    "Operation Short Text": "Op_Text",
                })
                df_ops["Planned_Hrs"] = pd.to_numeric(df_ops["Planned_Hrs"], errors="coerce").fillna(0)
                df_ops = df_ops[df_ops["Planned_Hrs"] > 0].copy()
            else:
                df_ops = None

        if df_comp_raw.empty:
            st.error(
                "**No components to process after filtering.**\n\n"
                "Likely causes:\n"
                "- All components already fully withdrawn (nothing left to pick)\n"
                "- All job start dates are more than 30 days in the past\n"
                "- COOIS export is missing material codes\n\n"
                "Check that your COOIS export covers current/upcoming jobs "
                "and includes the *Quantity withdrawn* column."
            )
            st.stop()

        with st.spinner("Running simulation..."):
            df_comp = simulate_picks(df_comp_raw, stock, sd_stock, log)
            df_comp = annotate_with_pos(df_comp, df_pos, log)
            df_comp = annotate_with_prs(df_comp, df_prs, log)
            df_jobs = aggregate_to_jobs(df_comp, log)
            df_jobs = annotate_subassembly_risk(df_jobs, df_header, log)

        with st.spinner("Building workbook..."):
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "READINESS_BOARD"
            build_readiness_board(ws1, df_jobs, today_str)
            build_component_detail(wb.create_sheet("COMPONENT_DETAIL"), df_comp, df_jobs, today_str)
            build_stock_ledger(wb.create_sheet("STOCK_LEDGER"), df_comp, stock, today_str)
            build_how_it_works(wb.create_sheet("HOW_IT_WORKS"))
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)

        st.session_state["df_jobs"] = df_jobs
        st.session_state["df_comp"] = df_comp
        st.session_state["df_ops"] = df_ops

        n_ready = int((df_jobs["Readiness"] == "READY").sum())
        n_partial = int((df_jobs["Readiness"] == "PARTIAL").sum())
        n_not = int((df_jobs["Readiness"] == "NOT READY").sum())

        st.success(f"Dashboard built — {today_str}")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("READY", n_ready, help="Safe to release to workshop")
        c2.metric("PARTIAL", n_partial, help="Small shortage — planner decides")
        c3.metric("NOT READY", n_not, help="Blocked — do not release")
        c4.metric("Total Jobs", len(df_jobs))

        stamp = datetime.now().strftime("%Y-%m-%d")
        st.download_button(
            label="Download Excel Dashboard",
            data=out,
            file_name=f"Job_Readiness_Board_{stamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

    except Exception as e:
        st.error(f"Build failed: {e}")

    if log_lines:
        with st.expander("Build log"):
            st.text("\n".join(log_lines))

elif not ready_to_run:
    if "df_jobs" not in st.session_state:
        st.info("Upload COOIS Components and MB52 Stock in the sidebar to enable the build.")


# ─── POST-BUILD TABS ─────────────────────────────────────────────────────────
if "df_jobs" in st.session_state:
    import plotly.express as px

    st.divider()

    df_jobs: pd.DataFrame = st.session_state["df_jobs"]
    df_comp: pd.DataFrame = st.session_state["df_comp"]
    _has_po = "PO_Doc" in df_comp.columns and df_comp["PO_Doc"].notna().any()
    _has_pr = "PR_Delivery_Date" in df_comp.columns and df_comp["PR_Delivery_Date"].notna().any()

    tab_board, tab_analytics, tab_load = st.tabs(["📋 Job Board", "📊 Analytics & Shortage Report", "🏭 Production Load"])

    # ── TAB 1: JOB BOARD ─────────────────────────────────────────────────────
    with tab_board:
        st.subheader("MO Lookup")

        search = st.text_input(
            "Search by MO number or job description",
            placeholder="e.g. 1234567 or pump",
        )

        if search:
            mask = (
                df_jobs["Order"].astype(str).str.contains(search, case=False, na=False)
                | df_jobs["Job_Desc"].astype(str).str.contains(search, case=False, na=False)
            )
            hits = df_jobs[mask].copy()
        else:
            hits = df_jobs.copy()

        _show_subassy = "Subassy_Risk" in hits.columns and (hits["Subassy_Risk"] != "-").any()
        _job_cols = [
            "Order", "Job_Desc", "SD_Order", "Start_Date", "Finish_Date", "Days_to_Start",
            "Readiness", "Components_Ready", "Components_Short", "Purchased_Short",
        ]
        if _show_subassy:
            _job_cols.append("Subassy_Risk")
        display = hits[_job_cols].copy()
        display["Start_Date"] = display["Start_Date"].dt.strftime("%d %b %Y")
        display["Finish_Date"] = display["Finish_Date"].dt.strftime("%d %b %Y")

        def _highlight_readiness(row):
            bg = READINESS_BG.get(row["Readiness"], "")
            return [
                f"background-color: {bg}; font-weight: bold" if col == "Readiness" else ""
                for col in row.index
            ]

        st.caption("Click a row to load its component breakdown below.")
        event = st.dataframe(
            display.style.apply(_highlight_readiness, axis=1),
            use_container_width=True,
            hide_index=True,
            selection_mode="single-row",
            on_select="rerun",
            column_config={
                "Order": st.column_config.TextColumn("MO Number", width="small"),
                "Job_Desc": st.column_config.TextColumn("Job Description"),
                "SD_Order": st.column_config.TextColumn("SD Order", width="small"),
                "Start_Date": st.column_config.TextColumn("Start Date", width="small"),
                "Finish_Date": st.column_config.TextColumn("Required Del. Date", width="small"),
                "Days_to_Start": st.column_config.NumberColumn("Days to Start", width="small"),
                "Readiness": st.column_config.TextColumn("Status", width="small"),
                "Components_Ready": st.column_config.NumberColumn("Parts Ready", width="small"),
                "Components_Short": st.column_config.NumberColumn("Parts Short", width="small"),
                "Purchased_Short": st.column_config.NumberColumn("Purchased Short", width="small"),
                "Subassy_Risk": st.column_config.TextColumn("Sub-assy Risk", width="medium"),
            },
        )
        st.caption(f"{len(hits)} job(s) shown")

        # Resolve selected MO — clicked row wins, otherwise default to first row
        _sel_rows = event.selection.rows
        if _sel_rows and _sel_rows[0] < len(hits):
            selected_mo = hits.iloc[_sel_rows[0]]["Order"]
        elif not hits.empty:
            selected_mo = hits.iloc[0]["Order"]
        else:
            selected_mo = None

        # ── Component breakdown ───────────────────────────────────────────────
        st.markdown("#### Component Breakdown")

        if selected_mo is not None:
            job_row = df_jobs[df_jobs["Order"] == selected_mo].iloc[0]
            readiness = job_row["Readiness"]
            bg = READINESS_BG.get(readiness, "#f0f0f0")

            st.markdown(
                f'<div style="background:{bg};padding:10px 18px;border-radius:6px;'
                f'font-weight:bold;font-size:15px;margin:8px 0 12px 0">'
                f'MO {selected_mo} &nbsp;|&nbsp; {job_row["Job_Desc"]} &nbsp;|&nbsp; {readiness}'
                f'</div>',
                unsafe_allow_html=True,
            )

            comps = df_comp[df_comp["Order"] == selected_mo].copy()

            comp_cols = [
                "Material", "Material_Desc", "Proc_Type",
                "Required", "Withdrawn", "To_Pick",
                "Stock_At_Turn", "Allocated", "Short_Qty",
                "Component_Status",
            ]
            comp_display = comps[comp_cols].copy()

            # ── Procurement columns (PO + PR) ─────────────────────────────
            job_start = job_row["Start_Date"]
            is_short = comp_display["Component_Status"] == "SHORT"

            if _has_po:
                comp_display["PO_Doc"] = comps["PO_Doc"].where(comps["PO_Doc"].notna(), "")
                comp_display["PO_Due"] = (
                    comps["PO_Delivery_Date"].dt.strftime("%d %b %Y")
                    .where(comps["PO_Delivery_Date"].notna(), "")
                )

            if _has_pr:
                comp_display["PR_Doc"] = comps["PR_Doc"].where(comps["PR_Doc"].notna(), "")
                comp_display["PR_Due"] = (
                    comps["PR_Delivery_Date"].dt.strftime("%d %b %Y")
                    .where(comps["PR_Delivery_Date"].notna(), "")
                )

            # Combined procurement status for SHORT components
            if _has_po or _has_pr:
                po_dates = comps["PO_Delivery_Date"] if _has_po else pd.Series(pd.NaT, index=comps.index)
                pr_dates = comps["PR_Delivery_Date"] if _has_pr else pd.Series(pd.NaT, index=comps.index)
                has_po = comps["PO_Doc"].notna() if _has_po else pd.Series(False, index=comps.index)
                has_pr = pr_dates.notna()

                proc_status = pd.Series("", index=comp_display.index)

                # PO exists — show PO timing
                po_on_time = is_short & has_po & (po_dates <= job_start)
                proc_status[po_on_time] = "PO: On Time"
                late_mask = is_short & has_po & (po_dates > job_start)
                if late_mask.any():
                    days_late = (po_dates[late_mask] - job_start).dt.days
                    proc_status[late_mask] = days_late.apply(lambda d: f"PO: Late ({d}d)")

                # No PO but PR exists
                pr_only_mask = is_short & ~has_po & has_pr
                proc_status[pr_only_mask] = "PR Only"

                # Nothing at all
                no_coverage = is_short & ~has_po & ~has_pr
                proc_status[no_coverage] = "No PR/PO"

                comp_display["Proc_Status"] = proc_status

            def _highlight_comp(row):
                bg_c = COMPONENT_BG.get(row["Component_Status"], "")
                result = []
                for col in row.index:
                    if col == "Component_Status":
                        result.append(f"background-color: {bg_c}; font-weight: bold")
                    elif col == "Proc_Status":
                        s = row["Proc_Status"]
                        if isinstance(s, str) and "Late" in s:
                            result.append("background-color: #FFC7CE; font-weight: bold")
                        elif s == "PO: On Time":
                            result.append("background-color: #C6EFCE")
                        elif s == "PR Only":
                            result.append("background-color: #FFEB9C")
                        elif s == "No PR/PO":
                            result.append("background-color: #FFC7CE")
                        else:
                            result.append("")
                    else:
                        result.append("")
                return result

            col_cfg = {
                "Material": st.column_config.TextColumn("Material", width="small"),
                "Material_Desc": st.column_config.TextColumn("Description"),
                "Proc_Type": st.column_config.TextColumn("Type", width="small"),
                "Required": st.column_config.NumberColumn("Required", format="%.2f", width="small"),
                "Withdrawn": st.column_config.NumberColumn("Withdrawn", format="%.2f", width="small"),
                "To_Pick": st.column_config.NumberColumn("To Pick", format="%.2f", width="small"),
                "Stock_At_Turn": st.column_config.NumberColumn("Stock Available", format="%.2f", width="small"),
                "Allocated": st.column_config.NumberColumn("Allocated", format="%.2f", width="small"),
                "Short_Qty": st.column_config.NumberColumn("Short Qty", format="%.2f", width="small"),
                "Component_Status": st.column_config.TextColumn("Status", width="small"),
            }
            if _has_po:
                col_cfg["PO_Doc"] = st.column_config.TextColumn("PO Number", width="small")
                col_cfg["PO_Due"] = st.column_config.TextColumn("PO Due Date", width="small")
            if _has_pr:
                col_cfg["PR_Doc"] = st.column_config.TextColumn("PR Number", width="small")
                col_cfg["PR_Due"] = st.column_config.TextColumn("PR Del. Date", width="small")
            if _has_po or _has_pr:
                col_cfg["Proc_Status"] = st.column_config.TextColumn("Procurement Status", width="medium")

            st.dataframe(
                comp_display.style.apply(_highlight_comp, axis=1),
                use_container_width=True,
                hide_index=True,
                column_config=col_cfg,
            )
        else:
            st.info("No jobs match your search.")

    # ── TAB 2: ANALYTICS & SHORTAGE REPORT ───────────────────────────────────
    with tab_analytics:
        today_ts = pd.Timestamp.today().normalize()

        n_ready   = int((df_jobs["Readiness"] == "READY").sum())
        n_partial = int((df_jobs["Readiness"] == "PARTIAL").sum())
        n_not     = int((df_jobs["Readiness"] == "NOT READY").sum())
        total_jobs = len(df_jobs)
        pct_ready = round(100 * n_ready / total_jobs) if total_jobs else 0
        total_short_lines = int((df_comp["Component_Status"] == "SHORT").sum())

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Ready", n_ready, f"{pct_ready}% of jobs")
        m2.metric("Partial", n_partial)
        m3.metric("Not Ready", n_not)
        m4.metric("Total Short Lines", total_short_lines)

        st.markdown("---")

        # ── Readiness donut + Jobs by week ────────────────────────────────────
        c_left, c_right = st.columns([1, 2])

        with c_left:
            st.markdown("##### Readiness Breakdown")
            fig_pie = px.pie(
                names=["READY", "PARTIAL", "NOT READY"],
                values=[n_ready, n_partial, n_not],
                color=["READY", "PARTIAL", "NOT READY"],
                color_discrete_map={
                    "READY": "#70AD47", "PARTIAL": "#FFC000", "NOT READY": "#FF4C4C"
                },
                hole=0.45,
            )
            fig_pie.update_traces(textinfo="value+percent", pull=[0.03, 0.03, 0.03])
            fig_pie.update_layout(margin=dict(t=20, b=20, l=20, r=20), height=280)
            st.plotly_chart(fig_pie, use_container_width=True)

        with c_right:
            st.markdown("##### Jobs by Required Delivery Week")
            jobs_plot = df_jobs.copy()
            jobs_plot["_Week"] = jobs_plot["Finish_Date"].dt.to_period("W").dt.start_time
            jobs_plot["Week"] = jobs_plot["_Week"].dt.strftime("%d %b")
            _week_order = (
                jobs_plot[["_Week", "Week"]].drop_duplicates()
                .sort_values("_Week")["Week"].tolist()
            )
            week_counts = (
                jobs_plot.groupby(["Week", "Readiness"])
                .size()
                .reset_index(name="Count")
            )
            fig_bar = px.bar(
                week_counts,
                x="Week",
                y="Count",
                color="Readiness",
                color_discrete_map={
                    "READY": "#70AD47", "PARTIAL": "#FFC000", "NOT READY": "#FF4C4C"
                },
                barmode="stack",
                category_orders={"Week": _week_order},
            )
            fig_bar.update_layout(
                margin=dict(t=20, b=40, l=20, r=20),
                height=280,
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
                xaxis_title="",
                yaxis_title="Jobs",
            )
            st.plotly_chart(fig_bar, use_container_width=True)

        st.markdown("---")

        # ── Delay risk analysis (requires PO data) ────────────────────────────
        st.markdown("##### Delay Risk Analysis")

        if _has_po:
            not_ready_orders = df_jobs[df_jobs["Readiness"] == "NOT READY"]["Order"].tolist()
            delay_rows = []
            for order in not_ready_orders:
                jr = df_jobs[df_jobs["Order"] == order].iloc[0]
                start = jr["Start_Date"]
                short_comps = df_comp[
                    (df_comp["Order"] == order) & (df_comp["Component_Status"] == "SHORT")
                ]
                with_po = short_comps[short_comps["PO_Delivery_Date"].notna()]
                no_po   = short_comps[short_comps["PO_Delivery_Date"].isna()]
                latest_po = with_po["PO_Delivery_Date"].max() if not with_po.empty else pd.NaT
                est_delay = (
                    int((latest_po - start).days)
                    if pd.notna(latest_po) and latest_po > start
                    else 0
                )
                delay_rows.append({
                    "Order": order,
                    "Job_Desc": jr["Job_Desc"],
                    "Start_Date": start.strftime("%d %b %Y"),
                    "Short_Parts": int(jr["Components_Short"]),
                    "Parts_No_PO": int(len(no_po)),
                    "Latest_PO_Due": latest_po.strftime("%d %b %Y") if pd.notna(latest_po) else "Unknown",
                    "Est_Delay_Days": est_delay,
                })

            if delay_rows:
                df_delay = pd.DataFrame(delay_rows).sort_values("Est_Delay_Days", ascending=False)
                total_delay = int(df_delay["Est_Delay_Days"].sum())
                uncovered_jobs = int((df_delay["Parts_No_PO"] > 0).sum())

                d1, d2, d3 = st.columns(3)
                d1.metric("NOT READY Jobs", len(df_delay))
                d2.metric("Total Est. Delay (job-days)", total_delay)
                d3.metric("Jobs with Uncovered Shorts", uncovered_jobs)

                st.dataframe(
                    df_delay,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Order": st.column_config.TextColumn("MO", width="small"),
                        "Job_Desc": st.column_config.TextColumn("Job Description"),
                        "Start_Date": st.column_config.TextColumn("Start Date", width="small"),
                        "Short_Parts": st.column_config.NumberColumn("Short Parts", width="small"),
                        "Parts_No_PO": st.column_config.NumberColumn("No PO Coverage", width="small"),
                        "Latest_PO_Due": st.column_config.TextColumn("Latest PO Due", width="small"),
                        "Est_Delay_Days": st.column_config.NumberColumn("Est. Delay (days)", width="small"),
                    },
                )

                top_delay = df_delay[df_delay["Est_Delay_Days"] > 0].head(15)
                if not top_delay.empty:
                    fig_delay = px.bar(
                        top_delay,
                        x="Est_Delay_Days",
                        y="Job_Desc",
                        orientation="h",
                        color="Est_Delay_Days",
                        color_continuous_scale=["#FFEB9C", "#FF4C4C"],
                        labels={"Est_Delay_Days": "Estimated Delay (days)", "Job_Desc": ""},
                    )
                    fig_delay.update_layout(
                        margin=dict(t=10, b=40, l=20, r=20),
                        height=max(250, len(top_delay) * 30),
                        coloraxis_showscale=False,
                    )
                    st.plotly_chart(fig_delay, use_container_width=True)
            else:
                st.info("No NOT READY jobs found.")
        else:
            st.info("Upload ZMPO purchase orders to see delay risk analysis.")

        st.markdown("---")

        # ── Next 2 weeks window ───────────────────────────────────────────────
        st.markdown("##### Next 2 Weeks — Jobs Starting Soon")

        two_weeks_out = today_ts + pd.Timedelta(days=14)
        df_2wk = df_jobs[
            (df_jobs["Start_Date"] >= today_ts) &
            (df_jobs["Start_Date"] <= two_weeks_out)
        ].sort_values("Start_Date").copy()

        if df_2wk.empty:
            st.info("No jobs scheduled to start in the next 14 days.")
        else:
            w_ready   = int((df_2wk["Readiness"] == "READY").sum())
            w_partial = int((df_2wk["Readiness"] == "PARTIAL").sum())
            w_not     = int((df_2wk["Readiness"] == "NOT READY").sum())

            w1, w2, w3, w4 = st.columns(4)
            w1.metric("Jobs in Window", len(df_2wk))
            w2.metric("READY", w_ready)
            w3.metric("PARTIAL", w_partial)
            w4.metric("NOT READY", w_not)

            _2wk_display_cols = ["Order", "Job_Desc", "SD_Order", "Start_Date", "Finish_Date",
                                  "Readiness", "Components_Short", "Purchased_Short"]
            if "Supply_Risk" in df_2wk.columns:
                _2wk_display_cols.append("Supply_Risk")
            _2wk_display_cols = [c for c in _2wk_display_cols if c in df_2wk.columns]
            disp_2wk = df_2wk[_2wk_display_cols].copy()
            disp_2wk["Start_Date"] = disp_2wk["Start_Date"].dt.strftime("%d %b %Y")
            disp_2wk["Finish_Date"] = disp_2wk["Finish_Date"].dt.strftime("%d %b %Y")

            st.dataframe(
                disp_2wk.style.apply(_highlight_readiness, axis=1),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Order": st.column_config.TextColumn("MO", width="small"),
                    "Job_Desc": st.column_config.TextColumn("Job Description"),
                    "SD_Order": st.column_config.TextColumn("SD Order", width="small"),
                    "Start_Date": st.column_config.TextColumn("Start Date", width="small"),
                    "Finish_Date": st.column_config.TextColumn("Required Del. Date", width="small"),
                    "Readiness": st.column_config.TextColumn("Status", width="small"),
                    "Components_Short": st.column_config.NumberColumn("Parts Short", width="small"),
                    "Purchased_Short": st.column_config.NumberColumn("Purchased Short", width="small"),
                    "Supply_Risk": st.column_config.TextColumn("Supply Risk", width="small"),
                },
            )

            # ── Export button ──────────────────────────────────────────────────
            _2wk_stamp = datetime.now().strftime("%Y-%m-%d")
            st.download_button(
                label="Export Next 2 Weeks to Excel",
                data=_build_2wk_excel(df_2wk, df_comp, today_str),
                file_name=f"Next_2_Weeks_{_2wk_stamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # ── Detailed shortage for 2-week jobs ─────────────────────────────
            st.markdown("###### Shortage Detail — Next 2 Weeks")
            orders_2wk = set(df_2wk["Order"].tolist())
            short_2wk = df_comp[
                df_comp["Order"].isin(orders_2wk) &
                (df_comp["Component_Status"] == "SHORT")
            ].copy()

            if short_2wk.empty:
                st.success("No shortages for jobs starting in the next 2 weeks.")
            else:
                short_2wk = short_2wk.sort_values(["Start_Date", "Order", "Material"])

                _s2_cols = ["Order", "Job_Desc", "Start_Date", "Material", "Material_Desc",
                            "Proc_Type", "Short_Qty"]
                _s2_cfg = {
                    "Order": st.column_config.TextColumn("MO", width="small"),
                    "Job_Desc": st.column_config.TextColumn("Job Description"),
                    "Start_Date": st.column_config.DateColumn("Start Date", format="DD MMM YYYY", width="small"),
                    "Material": st.column_config.TextColumn("Material", width="small"),
                    "Material_Desc": st.column_config.TextColumn("Description"),
                    "Proc_Type": st.column_config.TextColumn("Type", width="small"),
                    "Short_Qty": st.column_config.NumberColumn("Short Qty", format="%.2f", width="small"),
                }
                if _has_po and "PO_Delivery_Date" in short_2wk.columns:
                    short_2wk["PO_Due"] = (
                        short_2wk["PO_Delivery_Date"].dt.strftime("%d %b %Y")
                        .where(short_2wk["PO_Delivery_Date"].notna(), "—")
                    )
                    _s2_cols += ["PO_Due"]
                    _s2_cfg["PO_Due"] = st.column_config.TextColumn("PO Due", width="small")
                if _has_pr and "PR_Delivery_Date" in short_2wk.columns:
                    short_2wk["PR_Due"] = (
                        short_2wk["PR_Delivery_Date"].dt.strftime("%d %b %Y")
                        .where(short_2wk["PR_Delivery_Date"].notna(), "—")
                    )
                    _s2_cols += ["PR_Due"]
                    _s2_cfg["PR_Due"] = st.column_config.TextColumn("PR Due", width="small")

                def _highlight_proc(row):
                    result = []
                    for col in row.index:
                        if col == "PO_Due" and row.get("PO_Due", "—") not in ("—", ""):
                            result.append("background-color: #C6EFCE")
                        elif col == "PR_Due" and row.get("PR_Due", "—") not in ("—", ""):
                            result.append("background-color: #FFEB9C")
                        else:
                            result.append("")
                    return result

                st.dataframe(
                    short_2wk[_s2_cols].style.apply(_highlight_proc, axis=1),
                    use_container_width=True,
                    hide_index=True,
                    column_config=_s2_cfg,
                )
                st.caption(
                    f"{len(short_2wk)} short line(s) across "
                    f"{short_2wk['Order'].nunique()} job(s) starting in the next 14 days"
                )

        st.markdown("---")

        # ── Shortage report ───────────────────────────────────────────────────
        st.markdown("##### Shortage Report — All Jobs")

        short_all = df_comp[df_comp["Component_Status"] == "SHORT"].copy()

        if short_all.empty:
            st.success("No shortages across all jobs.")
        else:
            agg = {"Order": "nunique", "Short_Qty": "sum"}
            if _has_po:
                agg["PO_Delivery_Date"] = "max"
            if _has_pr:
                agg["PR_Delivery_Date"] = "min"

            shortage_summary = (
                short_all.groupby(["Material", "Material_Desc", "Proc_Type"])
                .agg(agg)
                .reset_index()
                .rename(columns={
                    "Order": "Jobs_Affected",
                    "Short_Qty": "Total_Short_Qty",
                    "PO_Delivery_Date": "Latest_PO_Due",
                    "PR_Delivery_Date": "Earliest_PR_Due",
                })
                .sort_values(["Jobs_Affected", "Total_Short_Qty"], ascending=False)
            )

            if _has_po:
                shortage_summary["Latest_PO_Due"] = (
                    shortage_summary["Latest_PO_Due"]
                    .dt.strftime("%d %b %Y")
                    .where(shortage_summary["Latest_PO_Due"].notna(), "—")
                )
            if _has_pr:
                shortage_summary["Earliest_PR_Due"] = (
                    shortage_summary["Earliest_PR_Due"]
                    .dt.strftime("%d %b %Y")
                    .where(shortage_summary["Earliest_PR_Due"].notna(), "—")
                )

            col_cfg_s = {
                "Material": st.column_config.TextColumn("Material", width="small"),
                "Material_Desc": st.column_config.TextColumn("Description"),
                "Proc_Type": st.column_config.TextColumn("Type", width="small"),
                "Jobs_Affected": st.column_config.NumberColumn("Jobs Affected", width="small"),
                "Total_Short_Qty": st.column_config.NumberColumn("Total Short Qty", format="%.2f", width="small"),
            }
            if _has_po:
                col_cfg_s["Latest_PO_Due"] = st.column_config.TextColumn("Latest PO Due", width="small")
            if _has_pr:
                col_cfg_s["Earliest_PR_Due"] = st.column_config.TextColumn("Earliest PR Due", width="small")

            st.dataframe(
                shortage_summary,
                use_container_width=True,
                hide_index=True,
                column_config=col_cfg_s,
            )
            st.caption(
                f"{len(shortage_summary)} unique material(s) short across "
                f"{int(short_all['Order'].nunique())} job(s)"
            )

            top_short = shortage_summary.head(12)
            fig_short = px.bar(
                top_short,
                x="Jobs_Affected",
                y="Material_Desc",
                orientation="h",
                color="Jobs_Affected",
                color_continuous_scale=["#FFEB9C", "#FF4C4C"],
                labels={"Jobs_Affected": "Jobs Affected", "Material_Desc": ""},
                title="Top Short Materials by Jobs Affected",
            )
            fig_short.update_layout(
                margin=dict(t=40, b=40, l=20, r=20),
                height=max(250, len(top_short) * 30),
                coloraxis_showscale=False,
            )
            st.plotly_chart(fig_short, use_container_width=True)

        # ── PO delivery vs job start scatter ──────────────────────────────────
        if _has_po:
            st.markdown("---")
            st.markdown("##### PO Delivery vs Job Start Date")

            po_risk = df_comp[
                df_comp["PO_Delivery_Date"].notna()
                & (df_comp["Component_Status"] == "SHORT")
            ].copy()

            if not po_risk.empty:
                # Map from df_jobs by Order to avoid merge column collisions
                # (df_comp already has Start_Date and Job_Desc so merging creates _x/_y suffixes)
                _job_idx = df_jobs.set_index("Order")
                po_risk["Job_Start"] = po_risk["Order"].map(_job_idx["Start_Date"])
                po_risk["Job_Name"] = po_risk["Order"].map(_job_idx["Job_Desc"])
                po_risk["Arrives Late"] = po_risk["PO_Delivery_Date"] > po_risk["Job_Start"]
                po_risk["Days Late"] = (
                    (po_risk["PO_Delivery_Date"] - po_risk["Job_Start"]).dt.days.clip(lower=0)
                )

                late_cnt    = int(po_risk["Arrives Late"].sum())
                on_time_cnt = int((~po_risk["Arrives Late"]).sum())
                p1, p2 = st.columns(2)
                p1.metric("PO Lines On Time", on_time_cnt)
                p2.metric("PO Lines Late (after job start)", late_cnt)

                all_dates = pd.concat([po_risk["Job_Start"], po_risk["PO_Delivery_Date"]])
                ref_min, ref_max = all_dates.min(), all_dates.max()

                fig_po = px.scatter(
                    po_risk.sort_values("Job_Start"),
                    x="Job_Start",
                    y="PO_Delivery_Date",
                    color="Arrives Late",
                    color_discrete_map={True: "#FF4C4C", False: "#70AD47"},
                    hover_data=["Material_Desc", "Job_Name", "Days Late"],
                    labels={
                        "Job_Start": "Job Start Date",
                        "PO_Delivery_Date": "PO Delivery Date",
                    },
                )
                fig_po.add_shape(
                    type="line",
                    x0=ref_min, y0=ref_min, x1=ref_max, y1=ref_max,
                    line=dict(color="gray", dash="dash", width=1),
                )
                fig_po.update_layout(margin=dict(t=20, b=40), height=400)
                st.plotly_chart(fig_po, use_container_width=True)
                st.caption("Points above the diagonal arrive after the job starts (at risk).")

    # ── TAB 3: PRODUCTION LOAD ────────────────────────────────────────────────
    with tab_load:
        df_ops = st.session_state.get("df_ops")

        if df_ops is None:
            st.info(
                "Upload a **COOIS Operations** file above and rebuild the dashboard "
                "to see production load.\n\n"
                "In COOIS: change the list type to **Operations**, apply the same filters "
                "as your Components export, and download to Excel."
            )
        else:
            # Map job fields onto operations
            _job_idx = df_jobs.set_index("Order")
            df_ops = df_ops.copy()
            df_ops["Readiness"] = df_ops["Order"].map(_job_idx["Readiness"])
            df_ops["Job_Desc"] = df_ops["Order"].map(_job_idx["Job_Desc"])
            df_ops["Job_Start"] = df_ops["Order"].map(_job_idx["Start_Date"])
            df_ops["Job_Finish"] = df_ops["Order"].map(_job_idx["Finish_Date"])

            # Spread each operation's hours evenly across every week the job spans.
            # This gives a much more accurate load profile than dumping all hours
            # into the job's start week.
            spread_rows = []
            for _, row in df_ops.iterrows():
                start = row["Job_Start"]
                if pd.isna(start):
                    continue
                finish = row["Job_Finish"]
                if pd.isna(finish) or finish < start:
                    finish = start
                start_w = start.to_period("W").start_time
                finish_w = finish.to_period("W").start_time
                weeks = pd.date_range(start_w, finish_w, freq="W-MON")
                if len(weeks) == 0:
                    weeks = pd.DatetimeIndex([start_w])
                hrs_per_wk = row["Planned_Hrs"] / len(weeks)
                base = row.to_dict()
                for w in weeks:
                    spread_rows.append({**base, "Week": w, "Week_Hrs": hrs_per_wk})

            if not spread_rows:
                st.warning("No operations with valid job dates found.")
                st.stop()

            df_spread = pd.DataFrame(spread_rows)
            df_spread["Week_Label"] = df_spread["Week"].dt.strftime("%d %b")

            # Ordered list of week labels sorted by actual date — fixes chart ordering
            week_order = (
                df_spread[["Week", "Week_Label"]]
                .drop_duplicates()
                .sort_values("Week")["Week_Label"]
                .tolist()
            )

            # Use Work_Centre_Desc if available, else fall back to Work_Centre code
            wc_col = "Work_Centre_Desc" if "Work_Centre_Desc" in df_spread.columns else "Work_Centre"

            # ── Filters ───────────────────────────────────────────────────────
            all_wcs = sorted(df_spread[wc_col].dropna().unique().tolist())
            f1, f2 = st.columns([2, 1])
            with f1:
                selected_wcs = st.multiselect("Filter work centres", all_wcs, default=all_wcs)
            with f2:
                capacity_hrs = st.number_input(
                    "Capacity per work centre per week (hours)", value=40, min_value=1
                )

            spread_filtered = df_spread[df_spread[wc_col].isin(selected_wcs)].copy()

            # ── Load by work centre per week ──────────────────────────────────
            st.markdown("##### Planned Hours by Work Centre per Week")
            st.caption("Hours are spread evenly across each job's planned start–finish span.")

            load_wc = (
                spread_filtered.groupby(["Week_Label", wc_col])["Week_Hrs"]
                .sum()
                .reset_index()
                .rename(columns={"Week_Label": "Week", wc_col: "Work Centre", "Week_Hrs": "Planned Hours"})
            )

            fig_load = px.bar(
                load_wc,
                x="Week",
                y="Planned Hours",
                color="Work Centre",
                barmode="stack",
                category_orders={"Week": week_order},
                color_discrete_sequence=px.colors.qualitative.Safe,
            )
            fig_load.add_hline(
                y=capacity_hrs,
                line_dash="dash",
                line_color="red",
                annotation_text=f"Capacity ({capacity_hrs}h)",
                annotation_position="top left",
            )
            fig_load.update_layout(
                margin=dict(t=20, b=40),
                height=360,
                xaxis_title="",
                yaxis_title="Planned Hours",
                legend_title="Work Centre",
            )
            st.plotly_chart(fig_load, use_container_width=True)

            # ── Overload table ────────────────────────────────────────────────
            overload = (
                spread_filtered.groupby(["Week_Label", wc_col])["Week_Hrs"]
                .sum()
                .reset_index()
                .rename(columns={"Week_Label": "Week", wc_col: "Work Centre", "Week_Hrs": "Planned Hours"})
            )
            overload = overload[overload["Planned Hours"] > capacity_hrs].sort_values(
                "Planned Hours", ascending=False
            )

            if not overload.empty:
                st.warning(
                    f"{len(overload)} work centre / week combination(s) exceed {capacity_hrs}h capacity."
                )
                st.dataframe(
                    overload,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Week": st.column_config.TextColumn("Week", width="small"),
                        "Work Centre": st.column_config.TextColumn("Work Centre"),
                        "Planned Hours": st.column_config.NumberColumn("Planned Hours", format="%.1f"),
                    },
                )
            else:
                st.success(f"No work centres exceed {capacity_hrs}h in any week.")

            st.markdown("---")

            # ── Load by readiness status per week ─────────────────────────────
            st.markdown("##### Load Split — Ready vs Not Ready Jobs")

            load_ready = (
                spread_filtered.groupby(["Week_Label", "Readiness"])["Week_Hrs"]
                .sum()
                .reset_index()
                .rename(columns={"Week_Label": "Week", "Week_Hrs": "Planned Hours"})
            )
            load_ready["Readiness"] = load_ready["Readiness"].fillna("Unknown")

            fig_ready = px.bar(
                load_ready,
                x="Week",
                y="Planned Hours",
                color="Readiness",
                barmode="stack",
                category_orders={"Week": week_order},
                color_discrete_map={
                    "READY": "#70AD47", "PARTIAL": "#FFC000",
                    "NOT READY": "#FF4C4C", "Unknown": "#BDBDBD",
                },
            )
            fig_ready.update_layout(
                margin=dict(t=20, b=40),
                height=300,
                xaxis_title="",
                yaxis_title="Planned Hours",
            )
            st.plotly_chart(fig_ready, use_container_width=True)
            st.caption("Hours in NOT READY jobs are at risk — parts are not available to release.")

            st.markdown("---")

            # ── Detailed operations table ─────────────────────────────────────
            with st.expander("Full operations list"):
                show_cols = ["Order", "Job_Desc", wc_col, "Op_Text", "Planned_Hrs", "Job_Start", "Job_Finish", "Readiness"]
                show_cols = [c for c in show_cols if c in df_spread.columns]
                detail = df_spread[show_cols].drop_duplicates().copy()
                for dc in ("Job_Start", "Job_Finish"):
                    if dc in detail.columns:
                        detail[dc] = detail[dc].dt.strftime("%d %b %Y")
                st.dataframe(detail, use_container_width=True, hide_index=True)
